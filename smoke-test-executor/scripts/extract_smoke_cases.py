#!/usr/bin/env python3
"""Extract normalized smoke cases from an Excel workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required to read .xlsx smoke case files") from exc


HEADER_ALIASES = {
    "caseId": {"测试编号", "用例编号", "caseId", "case_id", "id"},
    "title": {"标题", "用例标题", "名称", "title"},
    "modulePath": {"目录", "模块", "页面", "module", "path"},
    "precondition": {"前置条件", "precondition"},
    "steps": {"操作步骤", "步骤", "steps"},
    "expected": {"预期结果", "期望结果", "expected"},
    "priority": {"优先级", "priority"},
    "type": {"类型", "type"},
    "tags": {"标签", "tag", "tags"},
}


def split_lines(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    return [line.strip() for line in text.split("\n") if line.strip()]


def normalize_header(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    for normalized, aliases in HEADER_ALIASES.items():
        if text in aliases:
            return normalized
    return None


def load_cases(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    raw_headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    headers = [normalize_header(header) for header in raw_headers]
    cases: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        record: dict[str, Any] = {}
        for index, value in enumerate(row):
            if index >= len(headers) or headers[index] is None:
                continue
            record[headers[index]] = value
        if not any(record.values()):
            continue
        normalized = {
            "caseId": str(record.get("caseId") or "").strip(),
            "title": str(record.get("title") or "").strip(),
            "modulePath": str(record.get("modulePath") or "").strip(),
            "precondition": str(record.get("precondition") or "").strip(),
            "steps": split_lines(record.get("steps")),
            "expected": split_lines(record.get("expected")),
            "priority": str(record.get("priority") or "").strip(),
            "type": str(record.get("type") or "").strip(),
            "tags": str(record.get("tags") or "").strip(),
        }
        cases.append(normalized)
    return cases


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract normalized smoke cases from .xlsx")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet")
    parser.add_argument("--keyword", action="append", default=[])
    parser.add_argument("--priority", action="append", default=[])
    parser.add_argument("--tag", action="append", default=[])
    args = parser.parse_args()

    cases = load_cases(args.workbook, args.sheet)
    if args.keyword:
        needles = [keyword.lower() for keyword in args.keyword]
        cases = [
            case
            for case in cases
            if any(needle in json.dumps(case, ensure_ascii=False).lower() for needle in needles)
        ]
    if args.priority:
        priorities = set(args.priority)
        cases = [case for case in cases if case["priority"] in priorities]
    if args.tag:
        tags = args.tag
        cases = [case for case in cases if any(tag in case["tags"] for tag in tags)]

    print(json.dumps(cases, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
